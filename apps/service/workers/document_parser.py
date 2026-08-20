from __future__ import annotations

import re
from datetime import datetime
from io import BytesIO
from hashlib import sha256
from collections.abc import Callable, Iterator
from dataclasses import dataclass
from multiprocessing.synchronize import Event
from pathlib import Path
from zipfile import ZipFile
from xml.etree import ElementTree

import olefile
import xlrd
from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from pypdf import PdfReader

from domain.evidence import EvidenceLocator, ParseEvidence, ParseIssue, StructuredContentUnit


WORD_DOCUMENT_KINDS = frozenset({"doc", "docx", "docm", "dotx", "dotm"})
SPREADSHEET_DOCUMENT_KINDS = frozenset({"xls", "xlsx", "xlsm", "xltx", "xltm"})


class DocumentParseError(ValueError):
    """Raised when a local electronic document cannot be parsed safely."""


class DocumentParseCancelled(Exception):
    """Raised internally when a local parse is cancelled."""


@dataclass(frozen=True)
class DocumentPreflight:
    """A non-canonical inventory used to validate a converter's immutable input."""

    document_kind: str
    source_sha256: str
    inventory: dict[str, object]


def preflight_document(path: Path, document_kind: str) -> DocumentPreflight:
    """Inspect local source structure without producing Markdown or canonical evidence."""

    source_bytes = path.read_bytes()
    if document_kind == "pdf":
        return DocumentPreflight(document_kind, sha256(source_bytes).hexdigest(), _pdf_preflight(source_bytes))
    if document_kind in WORD_DOCUMENT_KINDS:
        inventory = _doc_preflight(source_bytes) if document_kind == "doc" else _docx_preflight(source_bytes)
        return DocumentPreflight(document_kind, sha256(source_bytes).hexdigest(), inventory)
    if document_kind in SPREADSHEET_DOCUMENT_KINDS:
        inventory = _xls_preflight(source_bytes) if document_kind == "xls" else _xlsx_preflight(source_bytes)
        return DocumentPreflight(document_kind, sha256(source_bytes).hexdigest(), inventory)
    raise DocumentParseError("This document format cannot be preflighted locally.")


def parse_items(
    items: tuple[dict[str, object], ...], should_cancel: Callable[[], bool] | None = None
) -> Iterator[dict[str, object]]:
    should_cancel = should_cancel or (lambda: False)
    yield {"type": "parse-started"}
    for item in items:
        if should_cancel():
            yield {"type": "parse-cancelled"}
            return
        path = Path(str(item["path"]))
        item_id = int(item["item_id"])
        try:
            source_bytes = path.read_bytes()
            content_sha256 = sha256(source_bytes).hexdigest()
            evidence = _parse_document_bytes(source_bytes, str(item["document_kind"]), should_cancel)
            if should_cancel():
                yield {"type": "parse-cancelled"}
                return
            yield {
                "type": "parse-item",
                "item_id": item_id,
                "content_sha256": content_sha256,
                "evidence": evidence.to_dict(),
            }
        except DocumentParseCancelled:
            yield {"type": "parse-cancelled"}
            return
        except DocumentParseError as error:
            yield {
                "type": "parse-failed-item",
                "item_id": item_id,
                "reason": str(error) or "The document could not be parsed.",
                "locator_summary": "document",
            }
        except OSError:
            yield {
                "type": "parse-failed-item",
                "item_id": item_id,
                "reason": "The source file is no longer available for local parsing.",
                "locator_summary": "document",
            }
    yield {"type": "parse-completed"}


def run_parse_worker(items: tuple[dict[str, object], ...], queue, cancelled: Event) -> None:
    for event in parse_items(items, cancelled.is_set):
        queue.put(event)


def parse_document(path: Path, document_kind: str) -> ParseEvidence:
    """Legacy V1 reader path retained for completed tasks, never a V2 converter."""
    return _parse_document_bytes(path.read_bytes(), document_kind)


def _pdf_preflight(source_bytes: bytes) -> dict[str, object]:
    try:
        reader = PdfReader(BytesIO(source_bytes))
        encrypted = reader.is_encrypted
        if encrypted and reader.decrypt("") == 0:
            return {
                "encrypted": True,
                "page_count": len(reader.pages),
                "text_pages": 0,
                "text_page_numbers": [],
            }
        text_page_numbers = [
            page_number
            for page_number, page in enumerate(reader.pages, start=1)
            if (page.extract_text() or "").strip()
        ]
        text_pages = len(text_page_numbers)
        return {
            "encrypted": False,
            "page_count": len(reader.pages),
            "text_pages": text_pages,
            "text_page_numbers": text_page_numbers,
            "text_coverage": text_pages / len(reader.pages) if reader.pages else 0.0,
        }
    except Exception as error:
        raise DocumentParseError("The PDF could not be preflighted.") from error


def _docx_preflight(source_bytes: bytes) -> dict[str, object]:
    try:
        with ZipFile(BytesIO(source_bytes)) as archive:
            document_xml = archive.read("word/document.xml")
            styles_xml = archive.read("word/styles.xml") if "word/styles.xml" in archive.namelist() else b""
    except Exception as error:
        raise DocumentParseError("The DOCX could not be preflighted.") from error
    root = ElementTree.fromstring(document_xml)
    namespace = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    math_namespace = "{http://schemas.openxmlformats.org/officeDocument/2006/math}"
    heading_levels = _docx_style_heading_levels(styles_xml, namespace)
    body = root.find(f"{namespace}body")
    if body is None:
        raise DocumentParseError("The DOCX main body is missing.")
    required_anchors: list[str] = []
    classifications: list[dict[str, object]] = []
    paragraph = table = 0
    for child in body:
        if child.tag == f"{namespace}p":
            paragraph += 1
            anchor = f"body/p[{paragraph}]"
            if not _docx_paragraph_has_content(child, namespace, math_namespace):
                continue
            required_anchors.append(anchor)
            heading_level = _docx_paragraph_heading_level(child, namespace, heading_levels)
            classification: dict[str, object] = {
                "anchor": anchor,
                "kind": "heading" if heading_level is not None else "paragraph",
            }
            if heading_level is not None:
                classification["heading_level"] = heading_level
            classifications.append(classification)
        elif child.tag == f"{namespace}tbl":
            table += 1
            anchor = f"body/tbl[{table}]"
            required_anchors.append(anchor)
            classifications.append({"anchor": anchor, "kind": "table"})
    return {
        "package_part_uri": "/word/document.xml",
        "required_anchors": required_anchors,
        "classifications": classifications,
    }


def _docx_style_heading_levels(styles_xml: bytes, namespace: str) -> dict[str, int]:
    if not styles_xml:
        return {}
    root = ElementTree.fromstring(styles_xml)
    heading_levels: dict[str, int] = {}
    for style in root.findall(f"{namespace}style"):
        style_id = style.get(f"{namespace}styleId")
        outline_level = style.find(f"{namespace}pPr/{namespace}outlineLvl")
        if not style_id or outline_level is None:
            continue
        try:
            level = int(outline_level.get(f"{namespace}val", "")) + 1
        except ValueError:
            continue
        if 1 <= level <= 6:
            heading_levels[style_id] = level
    return heading_levels


def _docx_paragraph_heading_level(
    paragraph: ElementTree.Element, namespace: str, style_heading_levels: dict[str, int]
) -> int | None:
    properties = paragraph.find(f"{namespace}pPr")
    if properties is None:
        return None
    outline_level = properties.find(f"{namespace}outlineLvl")
    if outline_level is not None:
        try:
            level = int(outline_level.get(f"{namespace}val", "")) + 1
        except ValueError:
            return None
        return level if 1 <= level <= 6 else None
    style = properties.find(f"{namespace}pStyle")
    style_id = style.get(f"{namespace}val") if style is not None else None
    return style_heading_levels.get(style_id or "")


def _docx_paragraph_has_content(
    paragraph: ElementTree.Element, namespace: str, math_namespace: str
) -> bool:
    """Ignore layout-only paragraphs that Pandoc correctly omits from its AST."""

    if "".join(paragraph.itertext()).strip():
        return True
    return any(
        paragraph.find(f".//{tag}") is not None
        for tag in (
            f"{namespace}drawing",
            f"{namespace}pict",
            f"{namespace}object",
            f"{math_namespace}oMath",
            f"{math_namespace}oMathPara",
        )
    )


def _parse_document_bytes(
    source_bytes: bytes,
    document_kind: str,
    should_cancel: Callable[[], bool] | None = None,
) -> ParseEvidence:
    should_cancel = should_cancel or (lambda: False)
    if should_cancel():
        raise DocumentParseCancelled
    if document_kind == "pdf":
        return _parse_pdf(source_bytes, should_cancel)
    if document_kind in WORD_DOCUMENT_KINDS:
        if document_kind == "doc":
            return _parse_doc(source_bytes, should_cancel)
        return _parse_docx(source_bytes, should_cancel, document_kind=document_kind)
    if document_kind in SPREADSHEET_DOCUMENT_KINDS:
        if document_kind == "xls":
            return _parse_xls(source_bytes, should_cancel)
        return _parse_xlsx(source_bytes, should_cancel, document_kind=document_kind)
    raise DocumentParseError("This document format cannot be parsed locally.")


def _parse_pdf(source_bytes: bytes, should_cancel: Callable[[], bool]) -> ParseEvidence:
    try:
        reader = PdfReader(BytesIO(source_bytes))
        if reader.is_encrypted and reader.decrypt("") == 0:
            raise DocumentParseError("The PDF is encrypted and cannot be read locally.")
    except DocumentParseError:
        raise
    except Exception as error:
        raise DocumentParseError("The PDF could not be read.") from error

    raw_pages: list[dict[str, object]] = []
    units: list[StructuredContentUnit] = []
    issues: list[ParseIssue] = []
    for page_number, page in enumerate(reader.pages, start=1):
        if should_cancel():
            raise DocumentParseCancelled
        locator = EvidenceLocator(page=page_number)
        try:
            text = page.extract_text() or ""
        except Exception as error:
            raw_pages.append({"page": page_number, "text": ""})
            issues.append(
                ParseIssue(
                    code="page-unreadable",
                    message=f"Page text extraction failed: {type(error).__name__}.",
                    locator=locator,
                )
            )
            continue
        raw_pages.append({"page": page_number, "text": text})
        page_units, page_issues = _units_from_lines(text.splitlines(), locator)
        units.extend(page_units)
        issues.extend(page_issues)

    if not raw_pages:
        issues.append(
            ParseIssue(
                code="missing-pages",
                message="The PDF has no readable pages.",
                locator=EvidenceLocator(region="document"),
            )
        )
    confidence = max(0.0, 0.95 - (0.25 * len(issues)))
    return ParseEvidence(
        document_kind="pdf",
        raw_extraction={"pages": raw_pages},
        units=tuple(units),
        confidence=confidence,
        issues=tuple(issues),
    )


def _parse_docx(
    source_bytes: bytes,
    should_cancel: Callable[[], bool],
    *,
    document_kind: str = "docx",
) -> ParseEvidence:
    try:
        document = Document(BytesIO(source_bytes))
    except Exception as error:
        raise DocumentParseError("The DOCX could not be read.") from error

    paragraphs: list[dict[str, str]] = []
    tables: list[dict[str, object]] = []
    body_items: list[tuple[str, object]] = []
    issues: list[ParseIssue] = []
    paragraph_index = 0
    table_index = 0
    for child in document.element.body.iterchildren():
        if should_cancel():
            raise DocumentParseCancelled
        if child.tag == qn("w:p"):
            paragraph_index += 1
            paragraph = Paragraph(child, document)
            location = f"paragraph:{paragraph_index}"
            style_name = getattr(paragraph.style, "name", None)
            if not isinstance(style_name, str):
                style_name = ""
            paragraphs.append({"location": location, "style": style_name, "text": paragraph.text})
            text = paragraph.text.strip()
            if text:
                body_items.append(
                    (
                        "paragraph",
                        StructuredContentUnit(
                            kind=_docx_paragraph_kind(style_name),
                            text=text,
                            locator=EvidenceLocator(docx_location=location),
                        ),
                    )
                )
            continue
        if child.tag != qn("w:tbl"):
            continue
        table_index += 1
        table = Table(child, document)
        rows: list[list[dict[str, str]]] = []
        table_units: list[StructuredContentUnit] = []
        for row_index, row in enumerate(table.rows, start=1):
            if should_cancel():
                raise DocumentParseCancelled
            cells: list[dict[str, str]] = []
            for cell_index, cell in enumerate(row.cells, start=1):
                if should_cancel():
                    raise DocumentParseCancelled
                location = f"table:{table_index}/row:{row_index}/cell:{cell_index}"
                text = cell.text.strip()
                cells.append({"location": location, "text": cell.text})
                if text:
                    table_units.append(
                        StructuredContentUnit(
                            kind="table-cell",
                            text=text,
                            locator=EvidenceLocator(docx_location=location),
                        )
                    )
            rows.append(cells)
        tables.append({"table": table_index, "rows": rows})

        body_items.append(("table", tuple(table_units)))

    units: list[StructuredContentUnit] = []
    index = 0
    while index < len(body_items):
        kind, value = body_items[index]
        if kind == "table":
            units.extend(value)
            index += 1
            continue
        unit = value
        if (
            _is_question(unit.text)
            and index + 1 < len(body_items)
            and body_items[index + 1][0] == "paragraph"
            and _is_answer(body_items[index + 1][1].text)
        ):
            units.append(
                StructuredContentUnit(
                    kind="question-answer",
                    text=f"{unit.text}\n{body_items[index + 1][1].text}",
                    locator=unit.locator,
                )
            )
            index += 2
            continue
        units.append(unit)
        index += 1

    if not units:
        issues.append(
            ParseIssue(
                code="empty-document",
                message="The Word document has no readable paragraphs or table cells.",
                locator=EvidenceLocator(docx_location="document"),
            )
        )
    confidence = max(0.0, 0.95 - (0.25 * len(issues)))
    return ParseEvidence(
        document_kind=document_kind,
        raw_extraction={"paragraphs": paragraphs, "tables": tables},
        units=tuple(units),
        confidence=confidence,
        issues=tuple(issues),
    )


def _doc_preflight(source_bytes: bytes) -> dict[str, object]:
    text = _legacy_doc_text(source_bytes)
    return {
        "format": "ole-word",
        "text_runs": len(text),
        "text_characters": sum(len(line) for line in text),
        "text_inventory_known": bool(text),
    }


def _parse_doc(source_bytes: bytes, should_cancel: Callable[[], bool]) -> ParseEvidence:
    if should_cancel():
        raise DocumentParseCancelled
    lines = _legacy_doc_text(source_bytes)
    locator = EvidenceLocator(region="document")
    units, issues = _units_from_lines(lines, locator) if lines else ([], [])
    if lines:
        issues.append(
            ParseIssue(
                code="legacy-word-review",
                message="Legacy Word text extraction needs review before it can be trusted.",
                locator=locator,
            )
        )
    else:
        issues.append(
            ParseIssue(
                code="empty-document",
                message="The legacy Word document has no readable text.",
                locator=locator,
            )
        )
    confidence = max(0.0, 0.85 - (0.25 * len(issues)))
    return ParseEvidence(
        document_kind="doc",
        raw_extraction={
            "paragraphs": [
                {"location": f"paragraph:{index}", "text": line}
                for index, line in enumerate(lines, start=1)
            ]
        },
        units=tuple(units),
        confidence=confidence,
        issues=tuple(issues),
    )


def _legacy_doc_text(source_bytes: bytes) -> list[str]:
    try:
        if not olefile.isOleFile(BytesIO(source_bytes)):
            raise DocumentParseError("The legacy Word file is not a valid OLE document.")
        with olefile.OleFileIO(BytesIO(source_bytes)) as document:
            if not document.exists("WordDocument"):
                raise DocumentParseError("The legacy Word document stream is missing.")
            stream = document.openstream("WordDocument").read()
    except DocumentParseError:
        raise
    except Exception as error:
        raise DocumentParseError("The legacy Word document could not be read locally.") from error

    candidates: list[tuple[int, str]] = []
    utf16_pattern = re.compile(rb"(?:(?:[\x09\x0a\x0d\x20-\x7e])\x00){4,}")
    for match in utf16_pattern.finditer(stream):
        candidates.append((match.start(), match.group().decode("utf-16le", errors="ignore")))
    if not candidates:
        ascii_pattern = re.compile(rb"[\x20-\x7e\x09\x0a\x0d]{4,}")
        candidates = [
            (match.start(), match.group().decode("cp1252", errors="ignore"))
            for match in ascii_pattern.finditer(stream)
        ]
    lines: list[str] = []
    for _, candidate in sorted(candidates):
        normalized = candidate.replace("\x0b", "\n").replace("\x0c", "\n")
        for line in normalized.splitlines():
            value = " ".join(line.split()).strip()
            if value and value not in lines:
                lines.append(value)
    return lines


def _xlsx_preflight(source_bytes: bytes) -> dict[str, object]:
    try:
        workbook = load_workbook(
            BytesIO(source_bytes), read_only=True, data_only=False, keep_links=False
        )
    except Exception as error:
        raise DocumentParseError("The Excel OOXML workbook could not be preflighted.") from error
    try:
        sheets = [
            {
                "name": sheet.title,
                "rows": int(sheet.max_row or 0),
                "columns": int(sheet.max_column or 0),
            }
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()
    return {"format": "ooxml-spreadsheet", "sheet_count": len(sheets), "sheets": sheets}


def _parse_xlsx(
    source_bytes: bytes,
    should_cancel: Callable[[], bool],
    *,
    document_kind: str,
) -> ParseEvidence:
    try:
        workbook = load_workbook(
            BytesIO(source_bytes), read_only=True, data_only=False, keep_links=False
        )
    except Exception as error:
        raise DocumentParseError("The Excel OOXML workbook could not be read.") from error
    sheets: list[dict[str, object]] = []
    units: list[StructuredContentUnit] = []
    issues: list[ParseIssue] = []
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            if should_cancel():
                raise DocumentParseCancelled
            rows: list[dict[str, object]] = []
            for row_index, row in enumerate(sheet.iter_rows(), start=1):
                if should_cancel():
                    raise DocumentParseCancelled
                cells: list[dict[str, str]] = []
                for cell in row:
                    if cell.value is None:
                        continue
                    text = _spreadsheet_value_text(cell.value)
                    if not text:
                        continue
                    location = f"sheet:{sheet.title}/cell:{cell.coordinate}"
                    cells.append({"location": location, "text": text})
                    units.append(
                        StructuredContentUnit(
                            kind="table-cell",
                            text=text,
                            locator=EvidenceLocator(region=location),
                        )
                    )
                if cells:
                    rows.append({"row": row_index, "cells": cells})
            sheets.append({"index": sheet_index, "name": sheet.title, "rows": rows})
    finally:
        workbook.close()
    if not units:
        issues.append(
            ParseIssue(
                code="empty-workbook",
                message="The Excel workbook has no non-empty cells.",
                locator=EvidenceLocator(region="workbook"),
            )
        )
    confidence = max(0.0, 0.95 - (0.25 * len(issues)))
    return ParseEvidence(
        document_kind=document_kind,
        raw_extraction={"sheets": sheets},
        units=tuple(units),
        confidence=confidence,
        issues=tuple(issues),
    )


def _xls_preflight(source_bytes: bytes) -> dict[str, object]:
    try:
        workbook = xlrd.open_workbook(file_contents=source_bytes, on_demand=True)
    except Exception as error:
        raise DocumentParseError("The legacy Excel workbook could not be preflighted.") from error
    try:
        sheets = [
            {"name": sheet.name, "rows": sheet.nrows, "columns": sheet.ncols}
            for sheet in workbook.sheets()
        ]
    finally:
        workbook.release_resources()
    return {"format": "biff-spreadsheet", "sheet_count": len(sheets), "sheets": sheets}


def _parse_xls(source_bytes: bytes, should_cancel: Callable[[], bool]) -> ParseEvidence:
    try:
        workbook = xlrd.open_workbook(file_contents=source_bytes, on_demand=True)
    except Exception as error:
        raise DocumentParseError("The legacy Excel workbook could not be read.") from error
    sheets: list[dict[str, object]] = []
    units: list[StructuredContentUnit] = []
    issues: list[ParseIssue] = []
    try:
        for sheet_index, sheet in enumerate(workbook.sheets(), start=1):
            rows: list[dict[str, object]] = []
            for row_index in range(sheet.nrows):
                if should_cancel():
                    raise DocumentParseCancelled
                cells: list[dict[str, str]] = []
                for column_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, column_index)
                    text = _xls_cell_text(cell, workbook.datemode)
                    if not text:
                        continue
                    coordinate = f"{_column_name(column_index)}{row_index + 1}"
                    location = f"sheet:{sheet.name}/cell:{coordinate}"
                    cells.append({"location": location, "text": text})
                    units.append(
                        StructuredContentUnit(
                            kind="table-cell",
                            text=text,
                            locator=EvidenceLocator(region=location),
                        )
                    )
                if cells:
                    rows.append({"row": row_index + 1, "cells": cells})
            sheets.append({"index": sheet_index, "name": sheet.name, "rows": rows})
    finally:
        workbook.release_resources()
    if not units:
        issues.append(
            ParseIssue(
                code="empty-workbook",
                message="The Excel workbook has no non-empty cells.",
                locator=EvidenceLocator(region="workbook"),
            )
        )
    confidence = max(0.0, 0.95 - (0.25 * len(issues)))
    return ParseEvidence(
        document_kind="xls",
        raw_extraction={"sheets": sheets},
        units=tuple(units),
        confidence=confidence,
        issues=tuple(issues),
    )


def _spreadsheet_value_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value).strip()


def _xls_cell_text(cell, datemode: int) -> str:
    if cell.ctype == xlrd.XL_CELL_EMPTY or cell.ctype == xlrd.XL_CELL_BLANK:
        return ""
    if cell.ctype == xlrd.XL_CELL_DATE:
        try:
            return xlrd.xldate_as_datetime(cell.value, datemode).isoformat(sep=" ")
        except (TypeError, ValueError, OverflowError):
            return str(cell.value).strip()
    if cell.ctype == xlrd.XL_CELL_ERROR:
        return f"#ERROR({cell.value})"
    return _spreadsheet_value_text(cell.value)


def _column_name(index: int) -> str:
    result = ""
    value = index + 1
    while value:
        value, remainder = divmod(value - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _units_from_lines(
    lines: list[str], locator: EvidenceLocator
) -> tuple[list[StructuredContentUnit], list[ParseIssue]]:
    units: list[StructuredContentUnit] = []
    issues: list[ParseIssue] = []
    normalized = [line.strip() for line in lines if line.strip()]
    if not normalized:
        return units, [
            ParseIssue(
                code="empty-page",
                message="No machine-readable text was extracted from this page.",
                locator=locator,
            )
        ]
    index = 0
    while index < len(normalized):
        line = normalized[index]
        if _is_question(line) and index + 1 < len(normalized) and _is_answer(normalized[index + 1]):
            units.append(
                StructuredContentUnit(
                    kind="question-answer",
                    text=f"{line}\n{normalized[index + 1]}",
                    locator=locator,
                )
            )
            index += 2
            continue
        units.append(StructuredContentUnit(kind=_pdf_line_kind(line, index), text=line, locator=locator))
        index += 1
    return units, issues


def _pdf_line_kind(text: str, index: int) -> str:
    if index == 0 and _looks_like_heading(text):
        return "heading"
    if re.match(r"^(?:[-*+]|[0-9]+[.)])\s+", text):
        return "list-item"
    if "\t" in text or " | " in text:
        return "table-row"
    return "paragraph"


def _docx_paragraph_kind(style_name: str) -> str:
    lowered = style_name.casefold()
    heading_match = re.search(r"heading\s*([0-9]+)", lowered)
    if heading_match:
        level = int(heading_match.group(1))
        return "heading" if level == 1 else f"heading-{level}"
    if "heading" in lowered:
        return "heading"
    if "list" in lowered:
        return "list-item"
    return "paragraph"


def _looks_like_heading(text: str) -> bool:
    return bool(re.match(r"^(?:chapter|unit|section|lesson)\b", text, re.IGNORECASE)) or (
        len(text) <= 90 and text.isupper()
    )


def _is_question(text: str) -> bool:
    return bool(re.match(r"^(?:q(?:uestion)?[.:]|[0-9]+[.)])\s*", text, re.IGNORECASE)) or text.endswith("?")


def _is_answer(text: str) -> bool:
    return bool(re.match(r"^(?:a(?:nswer)?[.:])\s*", text, re.IGNORECASE))
